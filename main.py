from datetime import timedelta

from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, \
    QPushButton, QMessageBox, QComboBox, QTableWidget, QTableWidgetItem, QListWidget, QDateEdit
from PyQt5.QtCore import Qt
from sqlalchemy import create_engine, ForeignKey, Column, String, Integer, MetaData, Table, select, Float, func, DateTime, Boolean
from sqlalchemy.orm import sessionmaker
from openpyxl import Workbook
from openpyxl.styles import Alignment


class Database:
    def __init__(self):
        engine = create_engine('postgresql://postgres:59723833@localhost/restaurant')
        metadata = MetaData()

        self.users = Table('users', metadata,
                           Column('id', Integer, primary_key=True),
                           Column('phone_number', String),
                           Column('username', String),
                           Column('password', String),
                           Column('is_admin', Boolean, default=False),
                           )

        self.categories = Table('categories', metadata,
                                Column('id', Integer, primary_key=True),
                                Column('name', String),
                                )

        self.dishes = Table('dishes', metadata,
                            Column('id', Integer, primary_key=True),
                            Column('name', String),
                            Column('category_id', Integer, ForeignKey('categories.id')),  # Внешний ключ на категории
                            Column('price', Float),
                            )

        self.orders = Table('orders', metadata,
                            Column('id', Integer, primary_key=True),
                            Column('user_id', Integer, ForeignKey('users.id')),
                            Column('created_date', DateTime, default=func.now()),
                            )

        self.order_details = Table('order_details', metadata,
                                   Column('id', Integer, primary_key=True),
                                   Column('order_id', Integer, ForeignKey('orders.id')),
                                   Column('dish_id', Integer, ForeignKey('dishes.id')),
                                   Column('quantity', Integer),
                                   Column('total_price', Float),
                                   )

        metadata.create_all(engine)
        Session = sessionmaker(bind=engine)
        self.session = Session()


class WelcomeWindow(QMainWindow):
    def __init__(self, db):
        super().__init__()

        self.db = db

        self.setWindowTitle('Ресторанная доставка')
        self.setGeometry(200, 200, 800, 400)

        layout = QVBoxLayout()

        self.label = QLabel('Добро пожаловать! Вы зарегистрированы?', self)
        layout.addWidget(self.label)

        self.register_button = QPushButton('Регистрация', self)
        self.register_button.clicked.connect(self.show_registration_window)
        layout.addWidget(self.register_button)

        self.login_button = QPushButton('Вход', self)
        self.login_button.clicked.connect(self.show_login_window)
        layout.addWidget(self.login_button)

        central_widget = QWidget(self)
        central_widget.setLayout(layout)
        self.setCentralWidget(central_widget)

    def show_registration_window(self):
        self.registration_window = RegistrationWindow(self.db)
        self.registration_window.show()

    def show_login_window(self):
        self.login_window = LoginWindow(self.db)
        self.login_window.show()

class RegistrationWindow(QWidget):
    def __init__(self, db):
        super().__init__()

        self.db = db

        layout = QVBoxLayout()

        self.phone_label = QLabel('Номер телефона:', self)
        self.phone_input = QLineEdit(self)
        layout.addWidget(self.phone_label)
        layout.addWidget(self.phone_input)

        self.username_label = QLabel('Логин:', self)
        self.username_input = QLineEdit(self)
        layout.addWidget(self.username_label)
        layout.addWidget(self.username_input)

        self.password_label = QLabel('Пароль:', self)
        self.password_input = QLineEdit(self)
        layout.addWidget(self.password_label)
        layout.addWidget(self.password_input)

        self.register_button = QPushButton('Зарегистрироваться', self)
        self.register_button.clicked.connect(self.register_user)
        layout.addWidget(self.register_button)

        self.setLayout(layout)

    def register_user(self):
        phone_number = self.phone_input.text()
        username = self.username_input.text()
        password = self.password_input.text()

        if phone_number and username and password:
            user_data = {
                'phone_number': phone_number,
                'username': username,
                'password': password,
            }

            self.db.session.execute(self.db.users.insert().values(user_data))
            self.db.session.commit()

            QMessageBox.information(self, 'Успех', 'Регистрация прошла успешно.')
            self.close()
        else:
            QMessageBox.warning(self, 'Ошибка', 'Заполните все поля.')


class LoginWindow(QWidget):
    def __init__(self, db):
        super().__init__()

        self.db = db

        layout = QVBoxLayout()

        self.username_label = QLabel('Логин:', self)
        self.username_input = QLineEdit(self)
        layout.addWidget(self.username_label)
        layout.addWidget(self.username_input)

        self.password_label = QLabel('Пароль:', self)
        self.password_input = QLineEdit(self)
        layout.addWidget(self.password_label)
        layout.addWidget(self.password_input)

        self.login_button = QPushButton('Войти', self)
        self.login_button.clicked.connect(self.login)
        layout.addWidget(self.login_button)

        self.setLayout(layout)

    def login(self):
        username = self.username_input.text()
        password = self.password_input.text()

        try:
            user = self.db.session.execute(select(self.db.users).where(self.db.users.c.username == username)).fetchone()

            if user and user.password == password:
                if user.is_admin:
                    # Открываем окно админ-панели
                    self.show_admin_panel()
                else:
                    self.current_user_id = user.id
                    self.close()
                    self.show_order_window()
                    print("Username:", username)
                    print("Password:", password)
                    print("User:", user)
            else:
                QMessageBox.warning(self, 'Ошибка', f'Неверные логин или пароль.')

        except Exception as e:
            QMessageBox.critical(self, 'Ошибка', f'Произошла ошибка: {str(e)}')
            print("An error occurred:", e)
            import traceback
            traceback.print_exc()

    def show_admin_panel(self):
        self.admin_panel = AdminPanel(self.db)
        self.admin_panel.show()

    def show_order_window(self):
        self.order_window = OrderWindow(self.db, self.current_user_id)
        self.order_window.show()

class AdminPanel(QWidget):
    def __init__(self, db):
        super().__init__()

        self.db = db

        layout = QVBoxLayout()

        # Добавьте элементы интерфейса для выбора периода
        self.start_date_label = QLabel('Начальная дата:', self)
        self.start_date_edit = QDateEdit(self)
        layout.addWidget(self.start_date_label)
        layout.addWidget(self.start_date_edit)

        self.end_date_label = QLabel('Конечная дата:', self)
        self.end_date_edit = QDateEdit(self)
        layout.addWidget(self.end_date_label)
        layout.addWidget(self.end_date_edit)

        # Кнопка запроса выгрузки в Excel
        self.export_button = QPushButton('Экспорт в Excel', self)
        self.export_button.clicked.connect(self.export_to_excel)
        layout.addWidget(self.export_button)

        self.setLayout(layout)

    def export_to_excel(self):
        start_date = self.start_date_edit.date().toPyDate()
        end_date = self.end_date_edit.date().toPyDate()

        print(f"Exporting orders from {start_date} to {end_date}")

        try:
            # Получите данные для выгрузки в Excel
            orders_data = self.db.session.query(
                self.db.orders.c.id,
                self.db.orders.c.user_id,
                self.db.users.c.phone_number,
                func.sum(self.db.order_details.c.total_price).label('total_price')
            ).join(
                self.db.order_details, self.db.orders.c.id == self.db.order_details.c.order_id
            ).join(
                self.db.users, self.db.orders.c.user_id == self.db.users.c.id  # Добавлен JOIN с таблицей users
            ).filter(
                self.db.orders.c.created_date.between(start_date, end_date + timedelta(days=1))
            ).group_by(
                self.db.orders.c.id, self.db.orders.c.user_id, self.db.users.c.phone_number
                # Включаем номер телефона в GROUP BY
            ).all()

            print("Orders data:", orders_data)

            # Создайте новую книгу Excel и получите активный лист
            workbook = Workbook()
            sheet = workbook.active

            # Заголовки столбцов
            sheet['A1'] = 'Order ID'
            sheet['B1'] = 'User ID'
            sheet['C1'] = 'Total Price'

            # Заполните данные из orders_data
            for row_index, order in enumerate(orders_data, start=2):
                sheet.cell(row=row_index, column=1, value=order.id)
                sheet.cell(row=row_index, column=2, value=order.user_id)
                sheet.cell(row=row_index, column=3, value=order.total_price)

            # Выравнивание текста в ячейках
            for column in sheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column[0].column_letter].width = adjusted_width

            # Сохраните книгу Excel
            workbook.save('orders_data.xlsx')

            QMessageBox.information(self, 'Успех', 'Выгрузка в Excel выполнена успешно.')
        except Exception as e:
            # Попробуем добавить вывод информации об ошибке для отладки
            print("An error occurred:", e)
            QMessageBox.critical(self, 'Ошибка', f'Произошла ошибка: {str(e)}')
            import traceback
            traceback.print_exc()

class OrderWindow(QMainWindow):  # Changed QWidget to QMainWindow
    def __init__(self, db, current_user_id):
        super().__init__()

        self.db = db
        self.current_user_id = current_user_id

        self.setWindowTitle('Ресторанная доставка')
        self.setGeometry(200, 200, 800, 400)

        layout = QVBoxLayout()

        self.category_label = QLabel('Выберите категорию:', self)
        self.category_combo = QComboBox(self)
        categories = self.db.session.execute(select(self.db.categories)).fetchall()
        for category in categories:
            self.category_combo.addItem(category.name)
        layout.addWidget(self.category_label)
        layout.addWidget(self.category_combo)

        self.dish_label = QLabel('Выберите блюдо:', self)
        self.dish_combo = QComboBox(self)
        layout.addWidget(self.dish_label)
        layout.addWidget(self.dish_combo)

        self.quantity_label = QLabel('Количество:', self)
        self.quantity_input = QLineEdit(self)
        layout.addWidget(self.quantity_label)
        layout.addWidget(self.quantity_input)

        self.add_to_cart_button = QPushButton('Добавить в корзину', self)
        self.add_to_cart_button.clicked.connect(self.add_to_cart)
        layout.addWidget(self.add_to_cart_button)

        # Добавим атрибут selected_items
        self.selected_items = []

        # Добавим QListWidget для отображения выбранных товаров
        self.cart_list = QListWidget(self)
        layout.addWidget(self.cart_list)

        central_widget = QWidget(self)
        central_widget.setLayout(layout)
        self.setCentralWidget(central_widget)

        # Кнопка "Просмотреть корзину"
        self.view_cart_button = QPushButton('Просмотреть корзину', self)
        self.view_cart_button.clicked.connect(self.view_cart)
        layout.addWidget(self.view_cart_button)

        # Connect the category change event to the update_dishes method
        self.category_combo.currentIndexChanged.connect(self.update_dishes)

        # Initialize dishes for the first time
        self.update_dishes()

        self.checkout_button = QPushButton('Оформить заказ', self)
        self.checkout_button.clicked.connect(self.checkout)
        layout.addWidget(self.checkout_button)

    def update_dishes(self):
        # Clear the current dish combo box
        self.dish_combo.clear()

        # Get the selected category name
        category_name = self.category_combo.currentText()

        # Retrieve dishes for the selected category
        category_id = self.db.session.execute(
            select(self.db.categories.c.id).where(self.db.categories.c.name == category_name)).scalar()

        dishes = self.db.session.execute(
            select(self.db.dishes).where(self.db.dishes.c.category_id == category_id)).fetchall()

        # Populate the dish combo box with the retrieved dishes
        for dish in dishes:
            self.dish_combo.addItem(f"{dish.name} - {dish.price:.2f} руб.", userData=dish.id)

    def add_to_cart(self):
        category_name = self.category_combo.currentText()
        dish_id = self.dish_combo.currentData()
        quantity = int(self.quantity_input.text())

        # Получить объект блюда из базы данных
        dish = self.db.session.execute(
            select(self.db.dishes).where(self.db.dishes.c.id == dish_id)).fetchone()

        if dish:
            # Получить цену блюда
            price = dish.price

            # Рассчитать общую стоимость
            total_price = quantity * price

            # Добавить выбранный товар в список для последующего просмотра
            item_text = f"{quantity} x {dish.name} ({category_name}) - {total_price:.2f} руб."
            self.selected_items.append(item_text)
            self.cart_list.addItem(item_text)

            QMessageBox.information(self, 'Успех', 'Блюдо добавлено в корзину.')
        else:
            QMessageBox.warning(self, 'Ошибка', f'Не удалось найти информацию о блюде: {dish_id}')

    def view_cart(self):
        cart_contents = '\n'.join(self.selected_items)
        if cart_contents:
            QMessageBox.information(self, 'Корзина', f'Ваш заказ:\n{cart_contents}')
        else:
            QMessageBox.warning(self, 'Корзина', 'Добавьте блюда в корзину.')

    def checkout(self):
        try:
            # Создаем новый заказ
            order_data = {
                'user_id': self.current_user_id,
                'created_date': func.now(),  # Устанавливаем текущую дату и время
            }
            result = self.db.session.execute(self.db.orders.insert().values(order_data))
            order_id = result.inserted_primary_key[0]

            # Добавляем детали заказа (блюда)
            for item_text in self.selected_items:
                # Разбираем информацию о блюде из текста
                parts = item_text.split(' ')
                quantity = int(parts[0])
                dish_name = parts[2]
                category_name = parts[3][1:-1]  # Убираем скобки
                total_price = float(parts[-2])

                # Получаем id блюда из базы данных
                dish_id = self.db.session.query(self.db.dishes.c.id).join(
                    self.db.categories, self.db.dishes.c.category_id == self.db.categories.c.id
                ).filter(
                    self.db.dishes.c.name == dish_name,
                    self.db.categories.c.name == category_name
                ).scalar()

                if dish_id:
                    # Добавляем информацию о блюде в таблицу order_details
                    order_detail_data = {
                        'order_id': order_id,
                        'dish_id': dish_id,
                        'quantity': quantity,
                        'total_price': total_price,
                    }
                    self.db.session.execute(self.db.order_details.insert().values(order_detail_data))

            # После успешного оформления заказа, вы можете очистить корзину и вывести сообщение
            self.selected_items = []
            self.cart_list.clear()

            # Фиксируем изменения в базе данных
            self.db.session.commit()
            QMessageBox.information(self, 'Успех', 'Заказ успешно оформлен!')
            self.close()

        except Exception as e:
            # Откатываем транзакцию в случае ошибки
            self.db.session.rollback()
            QMessageBox.critical(self, 'Ошибка', f'Произошла ошибка при оформлении заказа: {str(e)}')
            print("An error occurred:", e)
            import traceback
            traceback.print_exc()


if __name__ == '__main__':
    app = QApplication([])
    db = Database()
    welcome_window = WelcomeWindow(db)
    welcome_window.show()
    app.exec_()
