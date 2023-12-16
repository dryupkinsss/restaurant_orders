from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QLabel, QPushButton, QMessageBox, QDateEdit, QDialog, QLineEdit, \
    QTableWidget, QTableWidgetItem, QHeaderView
from openpyxl import Workbook
from sqlalchemy import func, distinct
from datetime import timedelta

class AdminPanel(QWidget):
    def __init__(self, db):
        super().__init__()

        self.db = db

        self.setWindowTitle('Панель администратора')
        self.setGeometry(200, 200, 800, 400)

        layout = QVBoxLayout()

        self.start_date_label = QLabel('Начальная дата:', self)
        self.start_date_edit = QDateEdit(self)
        layout.addWidget(self.start_date_label)
        layout.addWidget(self.start_date_edit)

        self.end_date_label = QLabel('Конечная дата:', self)
        self.end_date_edit = QDateEdit(self)
        layout.addWidget(self.end_date_label)
        layout.addWidget(self.end_date_edit)

        self.export_button = QPushButton('Экспорт в Excel', self)
        self.export_button.clicked.connect(self.export_to_excel)
        layout.addWidget(self.export_button)

        self.view_user_orders_button = QPushButton('Посмотреть заказы пользователя', self)
        self.view_user_orders_button.clicked.connect(self.show_view_user_orders_dialog)
        layout.addWidget(self.view_user_orders_button)

        self.setLayout(layout)

    def show_view_user_orders_dialog(self):
        dialog = UserOrdersDialog(self.db, self)
        dialog.exec_()

    def export_to_excel(self):
        start_date = self.start_date_edit.date().toPyDate()
        end_date = self.end_date_edit.date().toPyDate()

        print(f"Exporting orders from {start_date} to {end_date}")

        try:
            orders_data = self.db.session.query(
                self.db.users.c.id.label('user_id'),
                self.db.users.c.phone_number,
                func.count(distinct(self.db.orders.c.id)).label('order_count'),
                func.sum(self.db.order_details.c.total_price).label('total_price')
            ).outerjoin(
                self.db.orders, self.db.users.c.id == self.db.orders.c.user_id
            ).outerjoin(
                self.db.order_details, self.db.orders.c.id == self.db.order_details.c.order_id
            ).filter(
                self.db.orders.c.created_date.between(start_date, end_date + timedelta(days=1))
            ).group_by(
                self.db.users.c.id, self.db.users.c.phone_number
            ).all()

            print("Orders data:", orders_data)

            workbook = Workbook()
            sheet = workbook.active

            sheet['A1'] = 'ID пользователя'
            sheet['B1'] = 'Номер телефона'
            sheet['C1'] = 'Количество заказов'
            sheet['D1'] = 'Сумма всех заказов'

            for row_index, order in enumerate(orders_data, start=2):
                sheet.cell(row=row_index, column=1, value=order.user_id)
                sheet.cell(row=row_index, column=2, value=order.phone_number)
                sheet.cell(row=row_index, column=3, value=order.order_count)
                sheet.cell(row=row_index, column=4, value=order.total_price)

            for column in sheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column[0].column_letter].width = adjusted_width

            workbook.save('orders_data.xlsx')

            QMessageBox.information(self, 'Успех', 'Выгрузка в Excel выполнена успешно.')
        except Exception as e:
            print("An error occurred:", e)
            QMessageBox.critical(self, 'Ошибка', f'Произошла ошибка: {str(e)}')
            import traceback
            traceback.print_exc()


class UserOrdersDialog(QDialog):
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db

        self.setWindowTitle('Просмотр заказов пользователя')
        self.setGeometry(300, 300, 800, 400)

        layout = QVBoxLayout()

        self.phone_label = QLabel('Номер телефона пользователя:', self)
        self.phone_input = QLineEdit(self)
        layout.addWidget(self.phone_label)
        layout.addWidget(self.phone_input)

        self.view_orders_button = QPushButton('Посмотреть заказы', self)
        self.view_orders_button.clicked.connect(self.view_user_orders)
        layout.addWidget(self.view_orders_button)

        self.orders_table = QTableWidget(self)
        layout.addWidget(self.orders_table)

        self.setLayout(layout)

    def view_user_orders(self):
        phone_number = self.phone_input.text()

        try:
            user_orders = self.db.session.query(
                self.db.orders.c.id,
                self.db.orders.c.created_date,
                func.sum(self.db.order_details.c.total_price).label('total_price')
            ).join(
                self.db.order_details, self.db.orders.c.id == self.db.order_details.c.order_id
            ).join(
                self.db.users, self.db.orders.c.user_id == self.db.users.c.id
            ).filter(
                self.db.users.c.phone_number == phone_number
            ).group_by(
                self.db.orders.c.id, self.db.orders.c.created_date
            ).all()

            self.orders_table.setRowCount(len(user_orders))
            self.orders_table.setColumnCount(4)
            self.orders_table.setHorizontalHeaderLabels(['Номер заказа', 'Дата и время заказа', 'Сумма заказа', 'Блюда заказа'])

            for row_index, order in enumerate(user_orders):
                self.orders_table.setItem(row_index, 0, QTableWidgetItem(str(order.id)))
                self.orders_table.setItem(row_index, 1, QTableWidgetItem(str(order.created_date)))
                self.orders_table.setItem(row_index, 2, QTableWidgetItem(str(order.total_price)))

                # Добавляем блюда к ячейке 'Блюда заказа'
                dishes = self.db.session.query(
                    self.db.dishes.c.name,
                    self.db.order_details.c.quantity,
                    self.db.order_details.c.total_price
                ).join(
                    self.db.orders, self.db.orders.c.id == self.db.order_details.c.order_id
                ).join(
                    self.db.dishes, self.db.order_details.c.dish_id == self.db.dishes.c.id
                ).filter(
                    self.db.orders.c.id == order.id
                ).all()

                dishes_info = "\n".join(f"{dish.name} x{dish.quantity} - {dish.total_price}" for dish in dishes)
                item = QTableWidgetItem(dishes_info)
                item.setTextAlignment(Qt.AlignTop | Qt.AlignLeft)  # Выравнивание текста в ячейке
                self.orders_table.setItem(row_index, 3, item)

                # Задаем высоту строки в зависимости от количества блюд в заказе
                dishes_count = len(dishes)
                self.orders_table.setRowHeight(row_index, dishes_count * 20)  # Вы можете регулировать высоту

            self.orders_table.resizeColumnsToContents()
            self.orders_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)

        except Exception as e:
            QMessageBox.critical(self, 'Ошибка', f'Произошла ошибка: {str(e)}')
            import traceback
            traceback.print_exc()